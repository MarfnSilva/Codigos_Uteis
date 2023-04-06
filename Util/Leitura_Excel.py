# -*- coding: utf-8 # -*- 

# import encodings
# import pandas as pd

from GenericTrace import report_exception, trace
import requests, json,csv#,base64
from datetime import datetime, timedelta   
import time
import openpyxl

wb = openpyxl.load_workbook(filename ='Sample.xlsx')
ws = wb.active

for row in ws.iter_rows(min_row=2, max_col=10, max_row=10000, values_only=True):
    a,b,c,d,e,f,g,h,i,j = row
    if a != None:
        end_validity_dte = datetime.now() + timedelta(days=1825) # 1825 Days = 5 Years
        end_validity_str = end_validity_dte.strftime("%Y-%m-%dT%H:%M:%S")
        placa = []#row[3].replace('-', '') if row[3] != '' else None
        cpf = b.replace(".", "").replace("-", "").replace(" ", "") if b != '' else None                        
        user = {
            "FirstName" : a.upper().strip(),
            "IdNumber" : cpf, 
            "CHType" : int(i), 
            "CompanyID": h, 
            "PartitionID" : j, 
            "CHState" : 0, 
            "CHEndValidityDateTime" : end_validity_str
            }
        print(f'Importando {user["FirstName"]} - {user["IdNumber"]}')
        print(c)   
        #print(user)
        #time.sleep(0.3)

#print(openpyxl.__version__)
#page = ws['Página1']
#print(wb2.sheetnames)
# print(ws['c2'].value)
# for row in ws.values:
#     a,b,c = row
    
    # # user = {"FirstName" : a,
    # #         "IdNumber" : b,
    # #         "AuxText09" : c,
    # #         }
    # print(a)


